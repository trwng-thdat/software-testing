# -*- coding: utf-8 -*-
"""
HW06 - Dung file Excel test case + bang tong hop tu Main_Report.md va summary.json.
Chay: node/python; can openpyxl.
Xuat: hw6/testcases/HW06_TestCases_23127344.xlsx
"""
import io, os, re, json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
REP = os.path.join(ROOT, 'hw6', 'Main_Report.md')
SUM = os.path.join(ROOT, 'hw6', 'reports', 'summary.json')
OUTDIR = os.path.join(ROOT, 'hw6', 'testcases')
os.makedirs(OUTDIR, exist_ok=True)
OUT = os.path.join(OUTDIR, 'HW06_TestCases_23127344.xlsx')

md = io.open(REP, encoding='utf-8').read().split('\n')

def clean(c):
    c = c.strip()
    c = re.sub(r'\*\*(.+?)\*\*', r'\1', c)      # bo dam
    c = re.sub(r'~~(.+?)~~', r'\1', c)          # bo gach ngang
    c = c.replace('`', '')
    return c.strip()

def parse_tables():
    """Tra ve list (header[], rows[][]) cho moi bang markdown."""
    tables, i, n = [], 0, len(md)
    while i < n:
        if md[i].lstrip().startswith('|') and i+1 < n and re.match(r'^\s*\|[\s:|-]+\|\s*$', md[i+1]):
            header = [clean(x) for x in md[i].strip().strip('|').split('|')]
            rows, j = [], i+2
            while j < n and md[j].lstrip().startswith('|'):
                cells = [clean(x) for x in md[j].strip().strip('|').split('|')]
                rows.append(cells)
                j += 1
            tables.append((header, rows))
            i = j
        else:
            i += 1
    return tables

tables = parse_tables()

# ---------- Nhan kiem toan VALID/INVALID/INCOMPLETE tung TC (tu §4.2/§5.2/§6.2) ----------
def parse_verdicts():
    """Doc cac bang 'Chi tiet test case KHONG dat'; tra ve {id: (nhan, ly_do)}."""
    vmap = {}
    for header, rows in tables:
        # bang co cot dau la ID, cot 2 la Nhan (VALID/INVALID/INCOMPLETE)
        for r in rows:
            if len(r) < 4:
                continue
            id_cell, label, ai_wrote, reason = r[0], r[1], r[2], r[3]
            lab = label.strip().upper()
            if lab not in ('INVALID', 'INCOMPLETE', 'VALID'):
                continue
            # cot ID co the gom nhieu: "TC-API3-074 · -075 · -076 ..."
            ids = re.findall(r'TC-API[123]-\d{3}', id_cell)
            if not ids:
                continue
            # neu dong dang "TC-API3-074 · -075 ..." thi bat them cac hau to
            base = ids[0]
            prefix = base[:8]  # TC-API3-
            for suf in re.findall(r'[·,]\s*-?(\d{3})', id_cell):
                ids.append(prefix + suf)
            rs = reason.strip() or ('AI viet: ' + ai_wrote.strip())
            for i in set(ids):
                vmap[i] = (lab, rs)
    return vmap

VERDICTS = parse_verdicts()

TCID = re.compile(r'^(TC-API[123]-\d{3}[ab/]*|A[123]-E\d{2}|SPEC-BUG-\d{2})')

def find_tc_table(prefixes, min_rows=10):
    """Tim bang co nhieu dong bat dau bang test-case id thuoc prefixes."""
    best = None
    for header, rows in tables:
        tc_rows = [r for r in rows if r and TCID.match(r[0])]
        if len(tc_rows) >= min_rows and any(r[0].startswith(tuple(prefixes)) for r in tc_rows):
            if best is None or len(tc_rows) > len(best[1]):
                best = (header, tc_rows)
    return best

# ---------- styles ----------
HFILL = PatternFill('solid', fgColor='1e3a5f')
HFONT = Font(bold=True, color='ffffff', size=11)
TITLE = Font(bold=True, size=14, color='1e3a5f')
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
THIN = Border(*[Side(style='thin', color='cccccc')]*4)
ZEB = PatternFill('solid', fgColor='f2f6fb')

def style_header(ws, ncol, row=1):
    for c in range(1, ncol+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HFILL; cell.font = HFONT; cell.alignment = CENTER; cell.border = THIN

def write_sheet(wb, name, header, rows, widths):
    ws = wb.create_sheet(name)
    ws.append(header)
    style_header(ws, len(header))
    for k, r in enumerate(rows):
        r = (r + ['']*len(header))[:len(header)]
        ws.append(r)
        for c in range(1, len(header)+1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.alignment = WRAP; cell.border = THIN
            if k % 2 == 1:
                cell.fill = ZEB
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 28
    return ws

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ============ Sheet 1: Tong hop ============
S = json.load(io.open(SUM, encoding='utf-8'))
r = S['runs']; T = S['totals']
ws = wb.create_sheet('Tong hop')
ws['A1'] = 'HW06 - Kiem thu API - Bang tong hop test case'; ws['A1'].font = TITLE
ws['A2'] = 'MSSV 23127344 - Truong Thanh Dat - 23KTPM3'; ws['A2'].font = Font(italic=True, color='555555')
head = ['Lan chay', 'Test case', 'Iteration', 'HTTP call', 'Assertion', 'PASS', 'FAIL', 'Thoi gian (s)']
ws.append([]); ws.append(head)
hrow = ws.max_row
style_header(ws, len(head), hrow)
order = ['api1','api2','api3','data1','data2','data3','spec']
for key in order:
    d = r[key]
    ws.append([d['label'], d['testCases'], d['iterations'], d['requests'], d['assertions'],
               d['assertions']-d['assertionFailures'], d['assertionFailures'], round(d['durationMs']/1000,1)])
    for c in range(1,len(head)+1):
        ws.cell(row=ws.max_row, column=c).border = THIN
ws.append(['TONG', T['tcTotal'], '', T['requests'], T['assertions'], T['assertions']-T['failed'], T['failed'], ''])
for c in range(1,len(head)+1):
    cell = ws.cell(row=ws.max_row, column=c); cell.font = Font(bold=True); cell.border = THIN
# ghi chu
ws.append([]); ws.append(['Ghi chu: 22 assertion FAIL deu nam trong folder SPEC va CO Y DINH that bai (assertion viet theo dac ta).'])
ws.append(['3 folder API xanh tuyet doi. Chi tiet: hw6/reports/summary.md · bao cao HTML: hw6/reports/*.html'])
for w,ci in zip([46,11,10,10,11,8,8,13], range(1,9)):
    ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

# ============ Sheet 2-4: test case tung API ============
COLW = [13, 40, 14, 22, 34, 38, 14, 40]
HDR = ['ID', 'Tieu de', 'Ky thuat', 'Truy vet (Coverage/FR/SEC)', 'Input / Precondition',
       'Expected (status + body)', 'Nhan kiem toan', 'Ly do (nguoi ra soat)']
VALID_REASON = 'VALID - ky vong khop hanh vi probe that cua SUT; assertion PASS khi chay Newman (xem §4.4/§5.4/§6.4)'
def norm_rows(rows):
    """Chuan hoa 6 cot goc + gan them Nhan/Ly do tu VERDICTS (mac dinh VALID)."""
    out = []
    for r0 in rows:
        r0 = (r0 + ['']*6)[:6]
        tid = re.sub(r'[*~`]', '', r0[0]).strip()
        # id co the la "TC-API3-081a/b" (dong gop) -> lay dang chuan TC-APIx-ddd de tra
        m = re.search(r'TC-API[123]-\d{3}', tid)
        canon = m.group(0) if m else tid
        lab, reason = VERDICTS.get(tid) or VERDICTS.get(canon) or ('VALID', VALID_REASON)
        out.append(r0 + [lab, reason])
    return out
for api, pfx, sheet in (('API 1', ['TC-API1'], 'API1 - users_me (FR-04)'),
                        ('API 2', ['TC-API2'], 'API2 - order cancel (FR-10)'),
                        ('API 3', ['TC-API3'], 'API3 - coupons (FR-17)')):
    found = find_tc_table(pfx, min_rows=10)
    if not found:
        print('  ! khong tim thay bang', api); continue
    header, rows = found
    nr = norm_rows(rows)
    write_sheet(wb, sheet, HDR, nr, COLW)
    from collections import Counter
    cnt = Counter(x[6] for x in nr)
    print('  %s: %d TC | %s' % (sheet, len(rows), dict(cnt)))

# ============ Sheet 5: 15 TC tu bo sung (gop tu 3 bang §4.3/§5.3/§6.3) ============
EXTID = re.compile(r'^A[123]-E\d{2}$')
ext_rows = []
seen = set()
for header, rows in tables:
    for r0 in rows:
        if r0 and EXTID.match(r0[0]) and r0[0] not in seen:
            seen.add(r0[0])
            # cot: ID | Tieu de | Ky thuat/SEC | Input | Expected | Vi sao AI bo sot
            ext_rows.append((r0 + ['']*6)[:6])
ext_rows.sort(key=lambda x: x[0])
if ext_rows:
    write_sheet(wb, 'Mo rong (tu bo sung)',
                ['ID','Tieu de','Ky thuat / SEC','Input','Expected','Vi sao AI bo sot'],
                ext_rows, [10,44,20,40,40,44])
    print('  Mo rong: %d TC' % len(ext_rows))

# ============ Sheet 6: 16 bug + link issue ============
BUGS = [
 ("BUG-01",378,"User tu nang quyen len admin qua PUT /api/users/me","Critical","API 1","TC-API1-029","SEC-06"),
 ("BUG-02",379,"GET /api/users/me tra ve password plaintext","Critical","API 1","TC-API1-038","SEC-01"),
 ("BUG-03",380,"Token tu ky bang secret hardcode duoc chap nhan","Critical","API 1","A1-E03","SEC-02"),
 ("BUG-04",381,"phone khong duoc validate o backend","Minor","API 1","TC-API1-014","FR-04"),
 ("BUG-05",382,"GET /api/admin/users khong kiem role","Critical","API 1","A1-E04","SEC-03"),
 ("BUG-06",383,"User huy duoc don dang shipping","Major","API 2","TC-API2-018","FR-10"),
 ("BUG-07",384,"Admin dua don canceled ve delivered","Major","API 2","TC-API2-024","FR-10"),
 ("BUG-08",385,"GET /api/orders/:id khong yeu cau token","Critical","API 2","TC-API2-042","SEC-02"),
 ("BUG-09",386,"Thong bao loi huy don khong cho biet ly do","Minor","API 2","A2-E05","FR-10"),
 ("BUG-11",387,"User thuong tao duoc coupon","Critical","API 3","TC-API3-029","SEC-03"),
 ("BUG-12",388,"User thuong xoa duoc coupon (ke ca seed)","Critical","API 3","A3-E04","SEC-03"),
 ("BUG-13",389,"discount_value am lam TANG tien phai tra","Major","API 3","A3-E05","FR-17"),
 ("BUG-14",377,"max_uses_per_user chuoi '0' luu thanh so 0","Major","API 3","TC-API3-020","FR-17"),
 ("BUG-15",390,"code:null tao duoc nhieu lan, pha UNIQUE","Major","API 3","A3-E03","FR-17"),
 ("BUG-16",391,"Trung code tra 500 kem text driver SQLite","Minor","API 3","TC-API3-004","Xu ly loi"),
 ("BUG-17",392,"Thieu body tra 500 HTML thay vi 400 JSON","Minor","API 3","TC-API3-067","Xu ly loi"),
]
bug_rows = [[b, api, title, sev, fr, tc, '#%d'%num,
             'https://github.com/DuyITLOR/group05_eshop/issues/%d'%num] for b,num,title,sev,api,tc,fr in BUGS]
write_sheet(wb, 'Loi (16 bug)',
            ['ID','API','Tieu de','Muc do','FR/SEC','TC phat hien','Issue','Link'],
            bug_rows, [10,8,44,10,12,16,9,54])

wb.save(OUT)
print('\nDa xuat:', OUT)
print('So sheet:', len(wb.sheetnames), '->', wb.sheetnames)
